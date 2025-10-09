"""
Script de transformation des fichiers Grand Livre Comptes
Traduction de la logique TypeScript en Python
"""

import pandas as pd
import numpy as np
import openpyxl
import json
from datetime import datetime
from pathlib import Path


def format_date(date_str):
    """Convertit différents formats de date en DD/MM/YYYY"""
    if pd.isna(date_str) or date_str is None:
        return ""
    
    date_str = str(date_str).strip()
    
    # Si c'est vide
    if not date_str:
        return ""
    
    # Format: "2024-01-01 00:00:00" (datetime avec heure)
    if '-' in date_str and ':' in date_str:
        try:
            date_part = date_str.split(' ')[0]
            parts = date_part.split('-')
            if len(parts) == 3:
                year, month, day = parts
                # Valider que ce sont des nombres
                if year.isdigit() and month.isdigit() and day.isdigit():
                    return f"{day.zfill(2)}/{month.zfill(2)}/{year}"
        except:
            pass
    
    # Format: "2024-01-01" (date ISO)
    if '-' in date_str and date_str.count('-') == 2:
        try:
            parts = date_str.split('-')
            if len(parts) == 3 and all(p.isdigit() for p in parts):
                if len(parts[0]) == 4:  # YYYY-MM-DD
                    year, month, day = parts
                    return f"{day.zfill(2)}/{month.zfill(2)}/{year}"
                elif len(parts[2]) == 4:  # DD-MM-YYYY
                    day, month, year = parts
                    return f"{day.zfill(2)}/{month.zfill(2)}/{year}"
        except:
            pass
    
    # Format: "DD/MM/YYYY" ou "DD/MM/YY" (déjà formaté)
    if '/' in date_str:
        try:
            parts = date_str.split('/')
            if len(parts) == 3 and all(p.isdigit() for p in parts):
                day, month, year = parts
                if len(year) == 2:
                    year = f"20{year}"
                return f"{day.zfill(2)}/{month.zfill(2)}/{year}"
        except:
            pass
    
    # Format: DDMMYYYY (8 chiffres)
    if len(date_str) == 8 and date_str.isdigit():
        day = date_str[0:2]
        month = date_str[2:4]
        year = date_str[4:8]
        return f"{day}/{month}/{year}"
    
    # Format: DDMMYY (6 chiffres): 010124
    if len(date_str) == 6 and date_str.isdigit():
        day = date_str[0:2]
        month = date_str[2:4]
        year = date_str[4:6]
        return f"{day}/{month}/20{year}"
    
    # Format: DMMYY (5 chiffres): 10124 = 01/01/24
    if len(date_str) == 5 and date_str.isdigit():
        day = f"0{date_str[0]}"
        month = date_str[1:3]
        year = date_str[3:5]
        return f"{day}/{month}/20{year}"
    
    # Format: DMMY (4 chiffres): 1124 = 01/12/24
    if len(date_str) == 4 and date_str.isdigit():
        day = f"0{date_str[0]}"
        month = f"0{date_str[1]}"
        year = date_str[2:4]
        return f"{day}/{month}/20{year}"
    
    # Si aucun format reconnu, retourner tel quel
    return date_str


def parse_amount(value):
    """Parse un montant en float avec gestion robuste des formats"""
    if pd.isna(value) or value is None or value == '':
        return 0.0
    
    try:
        # Si c'est déjà un nombre
        if isinstance(value, (int, float)):
            return float(value)
        
        # Convertir en string et nettoyer
        value_str = str(value).strip()
        
        # Supprimer les espaces
        value_str = value_str.replace(' ', '')
        
        # Gérer les formats avec virgule comme séparateur décimal
        # Ex: "1.091.219,45" ou "1091,50"
        if ',' in value_str:
            # Supprimer tous les points (séparateurs de milliers)
            value_str = value_str.replace('.', '')
            # Remplacer virgule par point
            value_str = value_str.replace(',', '.')
        
        # Gérer le cas où il n'y a que des points (format US avec points de milliers)
        # Ex: "1.091.219" → "1091219"
        elif value_str.count('.') > 1:
            value_str = value_str.replace('.', '')
        
        # Supprimer les caractères non numériques sauf le point et le signe négatif
        clean_str = ''.join(c for c in value_str if c.isdigit() or c in ['.', '-'])
        
        if clean_str and clean_str not in ['.', '-', '-.']:
            return float(clean_str)
        return 0.0
        
    except Exception as e:
        return 0.0


def safe_str(value):
    """Convertit une valeur en string de manière sûre"""
    if pd.isna(value) or value is None:
        return ""
    return str(value).strip()


def is_numeric_string(value):
    """Vérifie si une valeur est une chaîne numérique OU une date"""
    if pd.isna(value) or value is None:
        return False
    s = str(value).strip()
    
    # Accepter les nombres purs
    if s.isdigit() and len(s) > 0:
        return True
    
    # Accepter les dates au format "YYYY-MM-DD" ou "YYYY-MM-DD HH:MM:SS"
    if '-' in s:
        date_part = s.split(' ')[0]  # Prendre la partie date
        parts = date_part.split('-')
        if len(parts) == 3 and all(p.isdigit() for p in parts):
            return True
    
    # Accepter les dates au format "DD/MM/YYYY"
    if '/' in s:
        parts = s.split('/')
        if len(parts) == 3 and all(p.isdigit() for p in parts):
            return True
    
    return False


def parse_grand_livre_comptes(file_path):
    """
    Parse un fichier Excel Grand Livre Comptes avec pandas pour meilleure gestion
    Retourne: dict avec comptesData et allTransactions
    """
    
    try:
        # Lire le fichier Excel avec pandas - MEILLEURE GESTION DES TYPES
        df = pd.read_excel(
            file_path,
            sheet_name=0,
            header=None,  # Pas d'en-tête automatique
            dtype=str,    # Tout lire comme string d'abord
            keep_default_na=False,  # Ne pas convertir automatiquement en NaN
            na_filter=False  # Ne pas filtrer les NA
        )
        
        # Remplacer les NaN par None
        df = df.replace({np.nan: None})
        
        # Convertir en liste de listes pour compatibilité avec le code existant
        data = df.values.tolist()
        
    except Exception as e:
        print(f"❌ Erreur lors de la lecture du fichier: {e}")
        # Fallback sur openpyxl
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
    
    comptes_data = []
    all_transactions = []
    
    # Variables pour métadonnées
    entite = ""
    date_gl = ""
    periode = ""
    
    # Variables pour le log de traitement
    traitement_log = {
        "fichier": Path(file_path).name,
        "date_traitement": datetime.now().isoformat(),
        "lignes_totales": len(data),
        "comptes_detectes": [],
        "lignes_ignorees": [],
        "erreurs": []
    }
    
    # Extraction des métadonnées depuis l'en-tête (10 premières lignes)
    for i in range(min(10, len(data))):
        row = data[i]
        
        # Chercher l'entité
        if not entite and row[0]:
            first_cell = safe_str(row[0])
            if first_cell and "Date" not in first_cell and "Impression" not in first_cell and "©" not in first_cell:
                entite = first_cell
        
        # Chercher la période
        if any("Période du" in safe_str(cell) for cell in row if cell):
            for j, cell in enumerate(row):
                if cell and "Période du" in safe_str(cell):
                    if i + 1 < len(data) and j + 1 < len(data[i + 1]):
                        date_fin = safe_str(data[i + 1][j + 1])
                        # Extraire la date au format DD/MM/YYYY
                        import re
                        match = re.search(r'(\d{2})/(\d{2})/(\d{2,4})', date_fin)
                        if match:
                            year = match.group(3)
                            if len(year) == 2:
                                year = f"20{year}"
                            periode = f"{year}{match.group(2)}"
                            date_gl = date_fin
                    break
    
    # Valeurs par défaut
    if not entite:
        entite = "ENVOL"
    if not periode:
        periode = "202412"
    if not date_gl:
        date_gl = "31/12/2024"
    
    traitement_log["entite"] = entite
    traitement_log["periode"] = periode
    traitement_log["date_gl"] = date_gl
    
    # Parser les comptes et transactions
    i = 0
    while i < len(data):
        row = data[i]
        
        if not row or len(row) == 0:
            i += 1
            continue
        
        col0 = safe_str(row[0])
        col1 = row[1] if len(row) > 1 else None
        col2 = safe_str(row[2]) if len(row) > 2 else ""
        
        # Détection d'un compte: col0 = 6 chiffres, col1 = None/vide, col2 = libellé
        is_compte = (col0 and 
                     is_numeric_string(col0) and 
                     len(col0) == 6 and 
                     (col1 is None or safe_str(col1) == "") and 
                     col2 and 
                     "Total" not in col2)
        
        if is_compte:
            compte = {
                "Numero_Compte": col0,
                "Libelle_Compte": col2,
                "Periode": periode,
                "Transactions": []
            }
            
            compte_log = {
                "numero": col0,
                "libelle": col2,
                "ligne_debut": i,
                "transactions_count": 0,
                "ligne_fin": i
            }
            
            # Chercher les transactions de ce compte
            j = i + 1
            while j < len(data):
                trans_row = data[j]
                
                if not trans_row or len(trans_row) == 0:
                    j += 1
                    continue
                
                trans_col0 = safe_str(trans_row[0]) if len(trans_row) > 0 else ""
                trans_col1 = trans_row[1] if len(trans_row) > 1 else None
                trans_col2 = safe_str(trans_row[2]) if len(trans_row) > 2 else ""
                
                # Stop si "Total compte"
                if "Total" in trans_col2:
                    compte_log["ligne_fin"] = j
                    break
                
                # Stop si nouveau compte
                is_new_compte = (trans_col0 and 
                                is_numeric_string(trans_col0) and 
                                len(trans_col0) == 6 and 
                                (trans_col1 is None or safe_str(trans_col1) == ""))
                
                if is_new_compte:
                    compte_log["ligne_fin"] = j - 1
                    break
                
                # C'est une transaction si: date (4-6 chiffres) + code journal présent
                # Dans la fonction parse_grand_livre_comptes, remplacer la partie détection transaction:

                # C'est une transaction si: date valide + code journal présent
                is_transaction = False
                if trans_col0 and trans_col1 is not None and safe_str(trans_col1) != "":
                    # Vérifier si col0 est une date (numérique ou format date)
                    if is_numeric_string(trans_col0):
                        trans_col0_clean = str(trans_col0).strip()
                        # Date numérique (4-6 chiffres) ou date formatée (avec - ou /)
                        if (trans_col0_clean.isdigit() and 4 <= len(trans_col0_clean) <= 8) or \
                        ('-' in trans_col0_clean) or ('/' in trans_col0_clean):
                            is_transaction = True
                
                if is_transaction:
                    try:
                        date = format_date(trans_col0)
                        code_journal = safe_str(trans_col1)
                        numero_piece = safe_str(trans_row[2]) if len(trans_row) > 2 else ""
                        libelle = safe_str(trans_row[5]) if len(trans_row) > 5 else ""
                        
                        # Les montants sont aux colonnes 11 (débit), 14 (crédit), 17 (solde)
                        debit = parse_amount(trans_row[11]) if len(trans_row) > 11 else 0.0
                        credit = parse_amount(trans_row[14]) if len(trans_row) > 14 else 0.0
                        solde = parse_amount(trans_row[17]) if len(trans_row) > 17 else 0.0
                        
                        transaction = {
                            "Date_GL": date_gl,
                            "Entite": entite,
                            "Compte": compte["Numero_Compte"],
                            "Date": date,
                            "Code_Journal": code_journal,
                            "Numero_Piece": numero_piece,
                            "Libelle_Ecriture": libelle,
                            "Debit": debit,
                            "Credit": credit,
                            "Solde": solde
                        }
                        
                        compte["Transactions"].append(transaction)
                        all_transactions.append(transaction)
                        compte_log["transactions_count"] += 1
                        
                    except Exception as e:
                        traitement_log["erreurs"].append({
                            "ligne": j,
                            "erreur": str(e),
                            "contenu": str(trans_row[:6])
                        })
                else:
                    # Ligne ignorée (ni compte ni transaction)
                    if trans_col0 or trans_col2:  # Si pas vide
                        raison = []
                        
                        # Analyser pourquoi la ligne est ignorée
                        if trans_col0 and is_numeric_string(trans_col0):
                            if len(trans_col0) >= 4 and len(trans_col0) <= 6:
                                if trans_col1 is None or safe_str(trans_col1) == "":
                                    raison.append("col0 est une date mais col1 (code journal) est vide")
                                else:
                                    raison.append("ressemble à une transaction mais validation échouée")
                            elif len(trans_col0) < 4:
                                raison.append(f"col0 a {len(trans_col0)} caractères (trop court pour une date)")
                            elif len(trans_col0) > 6:
                                raison.append(f"col0 a {len(trans_col0)} caractères (trop long pour une date)")
                        elif trans_col0 and not is_numeric_string(trans_col0):
                            raison.append("col0 n'est pas numérique")
                        elif not trans_col0 and trans_col2:
                            raison.append("col0 vide mais col2 remplie (ligne descriptive?)")
                        else:
                            raison.append("structure non reconnue")
                        
                        traitement_log["lignes_ignorees"].append({
                            "ligne": j,
                            "col0": trans_col0,
                            "col1": safe_str(trans_col1),
                            "col2": trans_col2,
                            "raison": " | ".join(raison)
                        })
                
                j += 1
            
            # Ajouter le compte s'il a des transactions
            if len(compte["Transactions"]) > 0:
                comptes_data.append(compte)
                traitement_log["comptes_detectes"].append(compte_log)
            else:
                # Compte sans transaction
                traitement_log["erreurs"].append({
                    "ligne": i,
                    "erreur": f"Compte {col0} sans transactions",
                    "contenu": f"{col0} - {col2}"
                })
            
            i = j
        else:
            i += 1
    
    # Statistiques finales
    traitement_log["statistiques"] = {
        "total_comptes": len(comptes_data),
        "total_transactions": len(all_transactions),
        "lignes_ignorees_count": len(traitement_log["lignes_ignorees"]),
        "erreurs_count": len(traitement_log["erreurs"])
    }
    
    return {
        "comptesData": comptes_data,
        "allTransactions": all_transactions,
        "traitementLog": traitement_log
    }


def save_outputs(result, output_dir, base_filename):
    """
    Sauvegarde les résultats en JSON et Excel
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    # Créer le dossier log s'il n'existe pas
    log_path = output_path / "log"
    log_path.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d")
    
    # 1. Sauvegarder le JSON structuré
    json_file = output_path / f"{base_filename}_{timestamp}.json"
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(result["comptesData"], f, ensure_ascii=False, indent=2)
    
    # 2. Sauvegarder l'Excel complet (transactions avec infos compte)
    flat_data = []
    for compte in result["comptesData"]:
        for trans in compte["Transactions"]:
            flat_data.append({
                "Numero_Compte": compte["Numero_Compte"],
                "Libelle_Compte": compte["Libelle_Compte"],
                "Periode": compte["Periode"],
                **trans
            })
    
    if flat_data:
        df = pd.DataFrame(flat_data)
        excel_file = output_path / f"{base_filename}_{timestamp}.xlsx"
        df.to_excel(excel_file, index=False, engine='openpyxl')
    
    # 3. Sauvegarder le log de traitement
    log_file = log_path / f"recap_{base_filename}_{timestamp}.json"
    with open(log_file, 'w', encoding='utf-8') as f:
        json.dump(result["traitementLog"], f, ensure_ascii=False, indent=2)
    
    return str(json_file), str(excel_file), str(log_file)


def transform_file(input_file, output_dir):
    """
    Fonction principale de transformation
    """
    print(f"🔄 Transformation de: {input_file}")
    
    # Parser le fichier
    result = parse_grand_livre_comptes(input_file)
    
    # Extraire le nom de base du fichier
    base_filename = Path(input_file).stem
    
    # Sauvegarder les résultats
    json_file, excel_file, log_file = save_outputs(result, output_dir, base_filename)
    
    # Afficher le résumé
    log = result["traitementLog"]
    stats = log["statistiques"]
    
    print(f"✅ Fichiers générés:")
    print(f"   - JSON: {json_file}")
    print(f"   - Excel: {excel_file}")
    print(f"   - Log: {log_file}")
    print(f"\n📊 Statistiques:")
    print(f"   - Comptes détectés: {stats['total_comptes']}")
    print(f"   - Transactions extraites: {stats['total_transactions']}")
    print(f"   - Lignes ignorées: {stats['lignes_ignorees_count']}")
    print(f"   - Erreurs: {stats['erreurs_count']}")
    
    if stats['erreurs_count'] > 0:
        print(f"\n⚠️  Consultez {log_file} pour les détails des erreurs")
    
    return json_file, excel_file, log_file


# Pour tester le script directement
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 3:
        print("Usage: python transform_excel.py <input_file> <output_dir>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_dir = sys.argv[2]
    
    transform_file(input_file, output_dir)